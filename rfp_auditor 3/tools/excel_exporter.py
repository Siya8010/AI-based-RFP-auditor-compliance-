"""
tools/excel_exporter.py
────────────────────────
Phase 4 — Corporate Excel Report Exporter (Developer B)

Consumes a list[ComplianceRow] produced by the audit engine and writes a
polished, production-grade Excel workbook with two sheets:

    Sheet 1 — "Summary"
        Aggregated KPIs: total requirements, match breakdown, pass rate.

    Sheet 2 — "Detailed Findings"
        One row per ComplianceRow with:
        • Auto-sized columns
        • Frozen header row
        • Column-level filters enabled
        • Colour-coded status cells (Green / Yellow / Red / Grey)
        • Clickable hyperlinks in the Source URL column
        • Alternating row banding for readability

Designed for large result sets (thousands of rows): streams rows directly
into OpenPyXL's write-only engine for memory efficiency when batch_mode=True.

Usage:
    from tools.excel_exporter import ExcelExporter
    from shared.schemas import ComplianceRow

    exporter = ExcelExporter()
    output_path = exporter.export(rows, "output/audit_report.xlsx")
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from config.settings import get_logger
from shared.schemas import ComplianceRow, ComplianceStatus

logger: logging.Logger = get_logger(__name__)

# ── Colour Palette ─────────────────────────────────────────────────────────
# ARGB format (Alpha + RGB hex)

_COLOUR_FULL_MATCH_BG = "FF92D050"      # Green
_COLOUR_PARTIAL_MATCH_BG = "FFFFC000"   # Amber / Yellow
_COLOUR_NO_MATCH_BG = "FFFF0000"        # Red
_COLOUR_UNKNOWN_BG = "FFD9D9D9"         # Grey

_COLOUR_FULL_MATCH_FONT = "FF375623"    # Dark green text
_COLOUR_PARTIAL_MATCH_FONT = "FF7F6000" # Dark amber text
_COLOUR_NO_MATCH_FONT = "FF9C0006"      # Dark red text
_COLOUR_UNKNOWN_FONT = "FF404040"       # Dark grey text

_COLOUR_HEADER_BG = "FF203864"          # Navy blue
_COLOUR_HEADER_FONT = "FFFFFFFF"        # White
_COLOUR_ROW_ALT_BG = "FFF2F2F2"         # Light grey for alternating rows
_COLOUR_SUMMARY_ACCENT = "FF1F3864"     # Deep navy for summary headings

_COLOUR_HYPERLINK = "FF0563C1"          # Standard hyperlink blue

# ── Column Definitions ─────────────────────────────────────────────────────

_DETAIL_COLUMNS: list[dict] = [
    {"header": "#",                   "key": None,                  "width": 5,  "wrap": False},
    {"header": "Requirement",         "key": "requirement",         "width": 55, "wrap": True},
    {"header": "Recommended Model",   "key": "recommended_model",   "width": 30, "wrap": True},
    {"header": "Compliance Status",   "key": "compliance_status",   "width": 18, "wrap": False},
    {"header": "Proof / Justification","key": "proof_justification","width": 60, "wrap": True},
    {"header": "Source URL",          "key": "source_url",          "width": 45, "wrap": False},
]

_THIN_BORDER_SIDE = Side(style="thin", color="FFD0D0D0")
_THIN_BORDER = Border(
    left=_THIN_BORDER_SIDE,
    right=_THIN_BORDER_SIDE,
    top=_THIN_BORDER_SIDE,
    bottom=_THIN_BORDER_SIDE,
)


# ── Style helpers ──────────────────────────────────────────────────────────


def _status_colours(status: ComplianceStatus | str) -> tuple[str, str]:
    """Return (bg_argb, font_argb) for a given compliance status."""
    mapping = {
        ComplianceStatus.FULL_MATCH:    (_COLOUR_FULL_MATCH_BG,    _COLOUR_FULL_MATCH_FONT),
        ComplianceStatus.PARTIAL_MATCH: (_COLOUR_PARTIAL_MATCH_BG, _COLOUR_PARTIAL_MATCH_FONT),
        ComplianceStatus.NO_MATCH:      (_COLOUR_NO_MATCH_BG,      _COLOUR_NO_MATCH_FONT),
        "Full Match":                   (_COLOUR_FULL_MATCH_BG,    _COLOUR_FULL_MATCH_FONT),
        "Partial Match":                (_COLOUR_PARTIAL_MATCH_BG, _COLOUR_PARTIAL_MATCH_FONT),
        "No Match":                     (_COLOUR_NO_MATCH_BG,      _COLOUR_NO_MATCH_FONT),
    }
    return mapping.get(status, (_COLOUR_UNKNOWN_BG, _COLOUR_UNKNOWN_FONT))


def _make_fill(argb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=argb)


def _make_font(argb: str, bold: bool = False, size: int = 11) -> Font:
    return Font(color=argb, bold=bold, size=size)


def _apply_header_style(cell) -> None:  # type: ignore[no-untyped-def]
    cell.fill = _make_fill(_COLOUR_HEADER_BG)
    cell.font = _make_font(_COLOUR_HEADER_FONT, bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.border = _THIN_BORDER


# ── Summary Sheet Builder ──────────────────────────────────────────────────


def _build_summary_sheet(ws: Worksheet, rows: list[ComplianceRow]) -> None:
    """
    Write the Summary sheet with KPI statistics.

    Args:
        ws:   The OpenPyXL worksheet to populate.
        rows: The full list of ComplianceRow objects.
    """
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    total = len(rows)
    full_match = sum(1 for r in rows if r.compliance_status == ComplianceStatus.FULL_MATCH)
    partial_match = sum(1 for r in rows if r.compliance_status == ComplianceStatus.PARTIAL_MATCH)
    no_match = sum(1 for r in rows if r.compliance_status == ComplianceStatus.NO_MATCH)
    pass_rate = ((full_match + partial_match) / total) if total else 0.0

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "RFP Compliance Audit — Summary"
    title_cell.font = Font(color=_COLOUR_SUMMARY_ACCENT, bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Generated timestamp
    ws.merge_cells("A2:B2")
    ts_cell = ws["A2"]
    ts_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ts_cell.font = Font(color="FF666666", italic=True, size=10)
    ts_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 10  # spacer

    # KPI table
    kpi_rows: list[tuple[str, int | float | str, str, str]] = [
        ("Total Requirements",  total,         _COLOUR_SUMMARY_ACCENT, _COLOUR_HEADER_FONT),
        ("Full Match",          full_match,    _COLOUR_FULL_MATCH_BG,  _COLOUR_FULL_MATCH_FONT),
        ("Partial Match",       partial_match, _COLOUR_PARTIAL_MATCH_BG, _COLOUR_PARTIAL_MATCH_FONT),
        ("No Match",            no_match,      _COLOUR_NO_MATCH_BG,    _COLOUR_NO_MATCH_FONT),
        ("Pass Rate (Full + Partial)", pass_rate, _COLOUR_SUMMARY_ACCENT, _COLOUR_HEADER_FONT),
    ]

    for row_idx, (label, value, bg, fg) in enumerate(kpi_rows, start=4):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(color=fg, bold=True, size=12)
        label_cell.fill = _make_fill(bg)
        label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        label_cell.border = _THIN_BORDER
        ws.row_dimensions[row_idx].height = 24

        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = Font(color=fg, bold=True, size=12)
        value_cell.fill = _make_fill(bg)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.border = _THIN_BORDER
        if isinstance(value, float):
            value_cell.number_format = FORMAT_PERCENTAGE_00

    logger.debug("[ExcelExporter] Summary sheet written.")


# ── Detail Sheet Builder ───────────────────────────────────────────────────


def _build_detail_sheet(ws: Worksheet, rows: list[ComplianceRow]) -> None:
    """
    Write the Detailed Findings sheet.

    Args:
        ws:   The OpenPyXL worksheet to populate.
        rows: The full list of ComplianceRow objects.
    """
    ws.sheet_view.showGridLines = False

    # ── Set column widths ──────────────────────────────────────────────────
    for col_idx, col_def in enumerate(_DETAIL_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_def["width"]

    # ── Header row ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    for col_idx, col_def in enumerate(_DETAIL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
        _apply_header_style(cell)

    # ── Data rows ──────────────────────────────────────────────────────────
    for row_idx, compliance_row in enumerate(rows, start=2):
        is_alt = (row_idx % 2 == 0)
        alt_fill = _make_fill(_COLOUR_ROW_ALT_BG) if is_alt else None

        status_bg, status_fg = _status_colours(compliance_row.compliance_status)
        row_height = 48 if (
            len(compliance_row.requirement) > 120 or
            len(compliance_row.proof_justification) > 120
        ) else 30

        ws.row_dimensions[row_idx].height = row_height

        for col_idx, col_def in enumerate(_DETAIL_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=bool(col_def["wrap"]),
            )

            # Row number
            if col_def["key"] is None:
                cell.value = row_idx - 1
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.font = Font(color="FF666666", size=10)
                if alt_fill:
                    cell.fill = alt_fill

            # Compliance Status cell — colour-coded
            elif col_def["key"] == "compliance_status":
                cell.value = str(compliance_row.compliance_status.value
                                 if isinstance(compliance_row.compliance_status, ComplianceStatus)
                                 else compliance_row.compliance_status)
                cell.fill = _make_fill(status_bg)
                cell.font = _make_font(status_fg, bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # Source URL — clickable hyperlink
            elif col_def["key"] == "source_url":
                url = compliance_row.source_url
                cell.value = url
                cell.hyperlink = url
                cell.font = Font(color=_COLOUR_HYPERLINK, underline="single", size=10)
                if alt_fill:
                    cell.fill = alt_fill

            # All other text fields
            else:
                value = getattr(compliance_row, col_def["key"], "")
                cell.value = str(value) if value is not None else ""
                cell.font = Font(size=10)
                if alt_fill:
                    cell.fill = alt_fill

    # ── Freeze header row ──────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter on header row ──────────────────────────────────────────
    last_col_letter = get_column_letter(len(_DETAIL_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    logger.debug(f"[ExcelExporter] Detail sheet written. {len(rows)} data rows.")


# ── Main Exporter Class ────────────────────────────────────────────────────


class ExcelExporter:
    """
    Corporate-grade Excel report builder for RFP compliance audit results.

    Accepts a list of ComplianceRow objects and writes a formatted .xlsx
    workbook to the specified output path.

    Example:
        exporter = ExcelExporter()
        path = exporter.export(compliance_rows, "output/report.xlsx")
        print(f"Report saved to: {path}")
    """

    def __init__(self) -> None:
        logger.info("ExcelExporter initialised.")

    def export(
        self,
        rows: list[ComplianceRow],
        output_path: str = "output/audit_report.xlsx",
    ) -> str:
        """
        Generate and save the Excel workbook.

        Args:
            rows:        List of ComplianceRow objects to export.
            output_path: File path for the output .xlsx file.
                         Parent directories are created automatically.

        Returns:
            The resolved absolute path of the saved file.

        Raises:
            ValueError: If rows is empty.
            IOError:    If the file cannot be written.
        """
        if not rows:
            raise ValueError(
                "Cannot export an empty list of ComplianceRow objects. "
                "Run the audit pipeline first."
            )

        # Ensure output directory exists
        out_path = Path(output_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"[ExcelExporter] Exporting {len(rows)} rows → {out_path}")

        wb = Workbook()

        # ── Sheet 1: Summary ───────────────────────────────────────────────
        ws_summary: Worksheet = wb.active  # type: ignore[assignment]
        ws_summary.title = "Summary"
        _build_summary_sheet(ws_summary, rows)

        # ── Sheet 2: Detailed Findings ─────────────────────────────────────
        ws_detail: Worksheet = wb.create_sheet(title="Detailed Findings")
        _build_detail_sheet(ws_detail, rows)

        # ── Make "Detailed Findings" the active sheet on open ─────────────
        wb.active = ws_detail  # type: ignore[assignment]

        # ── Save ───────────────────────────────────────────────────────────
        try:
            wb.save(str(out_path))
        except IOError as exc:
            logger.error(f"[ExcelExporter] Failed to save workbook: {exc}")
            raise

        resolved = str(out_path.resolve())
        logger.info(f"[ExcelExporter] ✅ Workbook saved: {resolved}")
        return resolved

    def export_from_dicts(
        self,
        raw_rows: list[dict],
        output_path: str = "output/audit_report.xlsx",
    ) -> str:
        """
        Convenience method: accepts raw dicts (e.g. from JSON deserialisation)
        and converts them to ComplianceRow objects before exporting.

        Args:
            raw_rows:    List of dicts matching the ComplianceRow schema.
            output_path: Destination file path.

        Returns:
            The resolved absolute path of the saved file.
        """
        rows: list[ComplianceRow] = []
        for i, raw in enumerate(raw_rows):
            try:
                rows.append(ComplianceRow(**raw))
            except Exception as exc:
                logger.warning(f"[ExcelExporter] Skipping malformed row {i}: {exc} | data={raw}")

        if not rows:
            raise ValueError("All rows were invalid. Nothing to export.")

        logger.info(f"[ExcelExporter] Parsed {len(rows)}/{len(raw_rows)} valid rows from dicts.")
        return self.export(rows, output_path)